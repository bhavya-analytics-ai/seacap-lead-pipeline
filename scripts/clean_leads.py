import re
import sys
import requests
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
import os

# ============================================================
# ENV & PATHS
# ============================================================
load_dotenv(Path(__file__).resolve().parent.parent / ".env")   # repo root

GOOGLE_API_KEY       = os.getenv("GOOGLE_MAPS_API_KEY", "").strip().strip('"')
MILLIONVERIFIER_KEY  = os.getenv("MILLIONVERIFIER_API_KEY", "").strip().strip('"')
COBALT_KEY           = os.getenv("COBALT_API_KEY", "").strip().strip('"')
TWILIO_SID           = os.getenv("TWILIO_ACCOUNT_SID", "").strip().strip('"')
TWILIO_TOKEN         = os.getenv("TWILIO_AUTH_TOKEN", "").strip().strip('"')

def _pick_input():
    if len(sys.argv) > 1:
        return sys.argv[1]
    incoming = Path(__file__).parent.parent / "incoming"
    files = sorted(
        [f for f in incoming.iterdir() if f.suffix.lower() in {".csv", ".xlsx", ".xls"}],
        key=lambda f: f.stat().st_mtime, reverse=True
    )
    if files:
        return str(files[0])
    print("❌ No file in incoming/ and no file passed as argument.")
    sys.exit(1)

INPUT_FILE = _pick_input()
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
STEM = Path(INPUT_FILE).stem

# Set to a number to only process that many rows (saves API credits while testing)
# Set to None to process the full file
TEST_MODE = None

# (Column detection is fully content-based — see score_column() inside main())

# ============================================================
# CLEANING HELPERS
# ============================================================
FAKE_NAMES = {"test","user","john doe","jane doe","na","n/a","none","unknown","asdf","admin","no name","sample"}
HONORIFICS = re.compile(r'^(mr\.?|mrs\.?|ms\.?|dr\.?|prof\.?|jr\.?|sr\.?)\s+', re.I)
COMPANY_INDICATORS = re.compile(r'\b(llc|inc|corp|ltd|co\.|company|enterprises|group|solutions|services|consulting)\b', re.I)
ILLEGAL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def strip_illegal(val):
    if pd.isna(val):
        return val
    return ILLEGAL_CHARS.sub('', str(val))

def clean_money(val):
    if pd.isna(val) or not str(val).strip():
        return None
    cleaned = re.sub(r'[^\d.]', '', str(val))
    try:
        return float(cleaned)
    except:
        return None

def clean_name(val):
    if pd.isna(val) or not str(val).strip():
        return ""
    name = HONORIFICS.sub('', str(val).strip()).strip().title()
    if any(f in name.lower() for f in FAKE_NAMES):
        return ""
    if re.search(r'\d', name):
        return ""
    if COMPANY_INDICATORS.search(name):
        return ""
    return name

def clean_phone(val):
    if pd.isna(val) or not str(val).strip():
        return ""
    raw = str(val).strip().split(".")[0]
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return ""

def clean_email(val):
    if pd.isna(val) or not str(val).strip():
        return ""
    email = str(val).strip().lower()
    if "@" in email and "." in email.split("@")[-1]:
        return email
    return ""

def clean_date(val):
    if pd.isna(val) or not str(val).strip():
        return ""
    try:
        return pd.to_datetime(val, errors='coerce').strftime('%Y-%m-%d')
    except:
        return ""

def split_full_name(val):
    if not val or pd.isna(val):
        return "", ""
    parts = str(val).strip().split()
    if len(parts) >= 2:
        return parts[0].title(), " ".join(parts[1:]).title()
    return str(val).strip().title(), ""

def extract_state(address):
    """Extract 2-letter US state from an address string."""
    if not address or pd.isna(address):
        return ""
    addr = str(address).strip()
    # State before zip: e.g. "Dallas, TX 75201"
    m = re.search(r'\b([A-Z]{2})\s+\d{5}', addr)
    if m:
        return m.group(1)
    # State after comma at end: e.g. "Dallas, TX"
    m = re.search(r',\s*([A-Z]{2})\s*$', addr)
    if m:
        return m.group(1)
    return ""

def _phone_col_score(colname):
    """Pre-score a phone column by its name before Twilio verification."""
    n = colname.lower()
    if "fax" in n:                                   return -5
    if any(x in n for x in ["mobile", "cell"]):      return 3
    if any(x in n for x in ["phone", "primary", "main"]): return 2
    if any(x in n for x in ["work", "office"]):      return 1
    return 0

def _email_col_score(colname, email_val=""):
    """Pre-score an email column by name + domain quality."""
    score = 0
    n = colname.lower()
    if any(x in n for x in ["primary", "email", "mail"]): score += 2
    if any(x in n for x in ["work", "personal"]):         score += 1
    if email_val and "@" in email_val:
        local  = email_val.split("@")[0].lower()
        domain = email_val.split("@")[-1].lower()
        if any(local.startswith(r) for r in ["info","sales","admin","contact","support","hello","office"]):
            score -= 2
        if domain and not any(d in domain for d in ["gmail","yahoo","hotmail","outlook","aol","icloud"]):
            score += 2
    return score

def _split_multi(val):
    """Split a cell that may contain multiple phone/email values separated by ; / |"""
    parts = re.split(r'[;/|]|\s{2,}', str(val).strip())
    return [p.strip() for p in parts if p.strip()]

def line_type_to_label(line_type):
    """Convert Twilio line_type to human-readable Cell / Landline / VoIP."""
    lt = str(line_type or "").lower()
    if any(w in lt for w in ["mobile", "cell"]):
        return "Cell"
    if any(w in lt for w in ["landline", "fixed"]):
        return "Landline"
    if "voip" in lt:
        return "VoIP"
    return lt.title() if lt else ""

# ============================================================
# VERIFICATION
# ============================================================
def verify_email_millionverifier(email):
    if not email:
        return "FORMAT-BAD"
    if not MILLIONVERIFIER_KEY:
        return "UNKNOWN"
    try:
        r = requests.get(
            "https://api.millionverifier.com/api/v3/",
            params={"api": MILLIONVERIFIER_KEY, "email": email},
            timeout=5
        )
        data = r.json()
        result = data.get("result", "").lower()
        quality = data.get("quality", "").lower()
        if result in ["ok", "valid"]:
            return "VERIFIED"
        elif result == "disposable" or quality == "disposable":
            return "DISPOSABLE"
        elif result == "spamtrap":
            return "SPAMTRAP"
        elif result in ["error", "invalid"]:
            return "INVALID"
        return "UNVERIFIED"
    except Exception:
        return "UNKNOWN"

def verify_business_google(company):
    if not company or not GOOGLE_API_KEY:
        return {"exists": None, "found_phone": "", "found_address": "", "status": "UNKNOWN"}
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/findplacefromtext/json",
            params={
                "input": company,
                "inputtype": "textquery",
                "fields": "name,formatted_phone_number,business_status,formatted_address",
                "key": GOOGLE_API_KEY
            },
            timeout=5
        )
        data = r.json()
        candidates = data.get("candidates", [])
        if not candidates:
            return {"exists": False, "found_phone": "", "found_address": "", "status": "NOT-FOUND"}
        place = candidates[0]
        biz_status = place.get("business_status", "").upper()
        exists = biz_status in ["OPERATIONAL", ""]
        return {
            "exists": exists,
            "found_phone": clean_phone(place.get("formatted_phone_number", "")),
            "found_address": place.get("formatted_address", ""),
            "status": "VERIFIED-ACTIVE" if exists else "VERIFIED-INACTIVE"
        }
    except Exception:
        return {"exists": None, "found_phone": "", "found_address": "", "status": "UNKNOWN"}

def verify_sos_cobalt(company):
    if not company or not COBALT_KEY:
        return "UNKNOWN"
    try:
        r = requests.get(
            "https://api.cobaltintelligence.com/v1/search",
            headers={"Authorization": f"Bearer {COBALT_KEY}"},
            params={"name": company},
            timeout=8
        )
        if r.status_code != 200:
            return "UNKNOWN"
        data = r.json()
        results = data.get("data", [])
        if not results:
            return "UNKNOWN"
        status = str(results[0].get("status", "")).lower()
        if "active" in status or "good standing" in status:
            return "active"
        elif "inactive" in status or "dissolved" in status or "revoked" in status:
            return "inactive"
        return "UNKNOWN"
    except Exception:
        return "UNKNOWN"

def verify_phone_twilio(phone):
    if not phone:
        return {"label": "FORMAT-BAD", "caller_name": "", "line_type": ""}
    if not TWILIO_SID or not TWILIO_TOKEN:
        return {"label": "UNKNOWN", "caller_name": "", "line_type": ""}
    try:
        digits = re.sub(r'\D', '', phone)
        if len(digits) == 10:
            digits = "1" + digits
        url = f"https://lookups.twilio.com/v2/PhoneNumbers/+{digits}"
        r = requests.get(
            url,
            params={"Fields": "line_type_intelligence,caller_name"},
            auth=(TWILIO_SID, TWILIO_TOKEN),
            timeout=8
        )
        if r.status_code == 404:
            return {"label": "FORMAT-BAD", "caller_name": "", "line_type": ""}
        if r.status_code != 200:
            return {"label": "UNKNOWN", "caller_name": "", "line_type": ""}
        data = r.json()
        line_info   = data.get("line_type_intelligence") or {}
        caller_info = data.get("caller_name") or {}
        line_type   = line_info.get("type", "").lower()
        caller_name = (caller_info.get("caller_name") or "").strip()
        valid = line_info.get("error_code") is None
        if not valid:
            return {"label": "FORMAT-BAD", "caller_name": "", "line_type": line_type}
        if line_type == "voip":
            return {"label": "VOIP", "caller_name": caller_name, "line_type": line_type}
        if caller_name:
            return {"label": "VERIFIED", "caller_name": caller_name, "line_type": line_type}
        return {"label": "VERIFIED-NO-NAME", "caller_name": "", "line_type": line_type}
    except Exception:
        return {"label": "UNKNOWN", "caller_name": "", "line_type": ""}

def check_name_phone_match(csv_name, twilio_name):
    if not twilio_name:
        return "NO-DATA"
    if not csv_name:
        return "UNKNOWN"
    csv_words    = set(csv_name.lower().split())
    twilio_words = set(twilio_name.lower().split())
    if csv_words & twilio_words:
        return "MATCH"
    return "MISMATCH"

def rank_phones(phone_list, csv_name):
    results = []
    for phone in phone_list:
        cleaned = clean_phone(phone)
        if not cleaned:
            results.append({
                "original": phone, "cleaned": str(phone),
                "label": "FORMAT-BAD", "caller_name": "", "line_type": "",
                "name_match": "UNKNOWN", "score": 0
            })
            continue
        twilio     = verify_phone_twilio(cleaned)
        label      = twilio["label"]
        caller_name = twilio["caller_name"]
        line_type  = twilio["line_type"]
        name_match = check_name_phone_match(csv_name, caller_name)
        if name_match == "MISMATCH":
            label = "MISMATCH"

        score = {"VERIFIED": 5, "VERIFIED-NO-NAME": 4, "SECONDARY": 3,
                 "VOIP": 2, "UNKNOWN": 1, "MISMATCH": 1, "FORMAT-BAD": 0}.get(label, 1)

        results.append({
            "original": phone, "cleaned": cleaned,
            "label": label, "caller_name": caller_name,
            "line_type": line_type, "name_match": name_match, "score": score
        })

    results.sort(key=lambda x: x["score"], reverse=True)
    return results

def rank_emails(email_list):
    results = []
    for email in email_list:
        cleaned = clean_email(email)
        if not cleaned:
            results.append({"original": email, "cleaned": str(email), "label": "FORMAT-BAD", "score": 0})
            continue
        label = verify_email_millionverifier(cleaned)
        score = {"VERIFIED": 5, "UNVERIFIED": 3, "UNKNOWN": 2,
                 "INVALID": 1, "DISPOSABLE": 1, "SPAMTRAP": 0, "FORMAT-BAD": 0}.get(label, 1)
        results.append({"original": email, "cleaned": cleaned, "label": label, "score": score})
    results.sort(key=lambda x: x["score"], reverse=True)
    return results

# ============================================================
# SORT INTO 4 LISTS
# Uses internal column names set during seacap build
# ============================================================
def _name_match(full_name_found, first, last, company):
    """Partial OR match — if any part of first, last, or company appears in Full Name Found."""
    if not full_name_found:
        return False
    found = full_name_found.lower()
    for part in [first, last]:
        if part and len(part) >= 2 and part.lower() in found:
            return True
    if company:
        # Check first word of company too
        company_words = [w for w in company.lower().split() if len(w) >= 3
                         and w not in ("llc", "inc", "ltd", "corp", "the", "and")]
        if any(w in found for w in company_words):
            return True
    return False

def _email_name_match(email, first, last, company):
    """Check if first, last, or company word appears in the email address (before @)."""
    if not email or "@" not in email:
        return False
    local = email.split("@")[0].lower()
    for part in [first, last]:
        if part and len(part) >= 2 and part.lower() in local:
            return True
    if company:
        company_words = [w for w in company.lower().split() if len(w) >= 3
                         and w not in ("llc", "inc", "ltd", "corp", "the", "and")]
        if any(w in local for w in company_words):
            return True
    return False

def get_list(row):
    if str(row.get("_dnc_flag", "")).upper() in ["Y", "YES", "1", "TRUE", "DO NOT CALL"]:
        return "DNC"
    _status_str   = str(row.get("_status", "")).lower()
    _status_words = set(re.findall(r'\b\w+\b', _status_str))
    if (_status_words & {"funded", "won"} or "closed won" in _status_str) \
            and "unfunded" not in _status_words and "not funded" not in _status_str:
        return "Funded"

    phone1_label    = str(row.get("Phone Status", "")).upper()
    email1_label    = str(row.get("Email Status", "")).upper()
    phone1          = str(row.get("Best Phone", "")).strip()
    email1          = str(row.get("Best Email", "")).strip()
    full_name_found = str(row.get("Full Name Found", "")).strip()
    first           = str(row.get("First Name", "")).strip()
    last            = str(row.get("Last Name", "")).strip()
    company         = str(row.get("Company", "")).strip()

    email_good        = email1_label == "VERIFIED" and bool(email1)
    email_salvageable = email1_label in ("INVALID", "UNVERIFIED") and \
                        _email_name_match(email1, first, last, company)
    name_match        = _name_match(full_name_found, first, last, company)

    # MISMATCH phone — always Qualified if Cell (can text regardless of email)
    if phone1_label == "MISMATCH":
        line_type = str(row.get("Cell/Landline", "")).upper()
        if line_type == "CELL":
            return "Qualified"
        if email_good or email_salvageable:
            return "Qualified"
        return "Needs Fixing"

    # VERIFIED-NO-NAME — Qualified if email good OR name/company in email address
    # Needs Fixing only if email invalid + no name match anywhere
    if phone1_label == "VERIFIED-NO-NAME":
        if email_good or email_salvageable:
            return "Qualified"
        if email1_label == "INVALID" and not full_name_found and not name_match:
            return "Needs Fixing"
        return "Qualified"

    # VERIFIED phone — always Qualified
    if phone1_label == "VERIFIED":
        return "Qualified"

    # VOIP — only Qualified if email is good (carrier-filtered for SMS, risky alone)
    if phone1_label == "VOIP":
        if email_good or email_salvageable:
            return "Qualified"
        return "Needs Fixing"

    # Fallback — any phone (not FORMAT-BAD, not VOIP) qualifies
    if phone1 and phone1_label not in ("FORMAT-BAD", "VOIP"):
        return "Qualified"
    if email_good or email_salvageable:
        return "Qualified"

    return "Needs Fixing"

def get_priority(row):
    """
    TEXT  = confirmed cell phone (VERIFIED or VERIFIED-NO-NAME).
    EMAIL = verified email wins over MISMATCH (wrong-person phone).
    MISMATCH cell used only as last resort when email is not verified.
    """
    phone_label  = str(row.get("Phone Status", "")).upper()
    email_label  = str(row.get("Email Status", "")).upper()
    line_type    = str(row.get("Cell/Landline", "")).upper()
    is_cell      = line_type == "CELL"
    email_strong = email_label in ("VERIFIED", "VALID", "OK")

    # Strong verified cell → TEXT (best for SMS/MCA)
    if phone_label in ("VERIFIED", "VERIFIED-NO-NAME") and is_cell:
        return "TEXT"
    # Verified email beats MISMATCH phone (don't text a stranger)
    if email_strong:
        return "EMAIL"
    # MISMATCH cell with no good email → TEXT (last resort, might still work)
    if phone_label == "MISMATCH" and is_cell:
        return "TEXT"
    # Fallback
    return "TEXT" if str(row.get("Best Phone", "")).strip() else "EMAIL"

def quality_score(row):
    """Score a row by data quality — higher = better lead."""
    phone_score = {"VERIFIED": 5, "VERIFIED-NO-NAME": 4, "SECONDARY": 3,
                   "VOIP": 2, "UNKNOWN": 1, "MISMATCH": 1, "FORMAT-BAD": 0, "": 0
                   }.get(str(row.get("Phone Status", "")), 1)
    email_score = {"VERIFIED": 5, "UNVERIFIED": 3, "UNKNOWN": 2,
                   "INVALID": 1, "DISPOSABLE": 1, "SPAMTRAP": 0, "FORMAT-BAD": 0, "": 0
                   }.get(str(row.get("Email Status", "")), 1)
    return phone_score * 2 + email_score  # phone weighted higher

# ============================================================
# FIXED OUTPUT COLUMN ORDER (Adam's uniform template)
# ============================================================
FIXED_COLS = [
    "First Name", "Last Name", "Company",
    "Best Phone", "Phone Status", "Cell/Landline",
    "Best Email", "Email Status", "State",
    "Annual Revenue", "Flag Reason", "Full Name Found",
    "Fix_Reason", "List", "Duplicate_Flag",
    "Status", "Adam List", "Priority", "Notes",
]

# Close CRM status + Adam List number per list type
_LIST_META = {
    "Qualified":    ("🆕 New / Uncontacted",   "List 1 - Qualified"),
    "Needs Fixing": ("⚠️ Invalid Contact Info", "List 2 - Needs Fixing"),
    "DNC":          ("🔴 Not Interested - DNC", "List 3 - DNC"),
    "Funded":       ("🟢 Funded - SEACAP",      "List 4 - Funded"),
}

# ============================================================
# MAIN
# ============================================================
def main():
    print(f"\n📂 Loading: {INPUT_FILE}")

    def looks_like_data_row(row):
        vals = [str(v).strip() for v in row if str(v).strip() not in ("", "nan")]
        if not vals:
            return False
        has_email  = any("@" in v for v in vals)
        has_phone  = any(re.sub(r'\D','',v).__len__() >= 9 for v in vals)
        has_number = any(re.match(r'^\d[\d,\.]+$', v) for v in vals)
        return has_email or has_phone or has_number

    def load_file(fp):
        fp = str(fp)
        if fp.endswith((".xlsx", ".xls")):
            df = pd.read_excel(fp, header=0)
            cols = [str(c).strip() for c in df.columns]
            if looks_like_data_row(cols):
                df = pd.read_excel(fp, header=None)
                df.columns = [f"Col_{i}" for i in range(len(df.columns))]
            elif all(c.lstrip('-').isdigit() for c in cols):
                df = pd.read_excel(fp, header=None)
                for idx in range(min(10, len(df))):
                    row = df.iloc[idx]
                    str_count = sum(1 for v in row if isinstance(v, str) and len(str(v).strip()) > 1)
                    if str_count >= len(df.columns) * 0.5 and not looks_like_data_row(row):
                        df.columns = [str(v).strip() for v in df.iloc[idx]]
                        df = df.iloc[idx+1:].reset_index(drop=True)
                        break
                else:
                    df.columns = [f"Col_{i}" for i in range(len(df.columns))]
            return df
        elif fp.endswith(".csv"):
            for enc in ["utf-8", "latin1", "cp1252"]:
                try:
                    df = pd.read_csv(fp, low_memory=False, encoding=enc)
                    if all(str(c).strip().lstrip('-').isdigit() for c in df.columns):
                        for skip in range(1, 10):
                            df2 = pd.read_csv(fp, low_memory=False, encoding=enc, skiprows=skip)
                            if not all(str(c).strip().lstrip('-').isdigit() for c in df2.columns):
                                df = df2
                                break
                    return df
                except UnicodeDecodeError:
                    continue
            raise ValueError("Could not decode CSV with any known encoding")
        else:
            print("❌ Unsupported file type.")
            sys.exit(1)

    try:
        df = load_file(INPUT_FILE)
    except FileNotFoundError:
        print(f"❌ Not found: {INPUT_FILE}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Failed to load file: {e}")
        sys.exit(1)

    df.columns = [str(c).strip() for c in df.columns]
    # Dedupe column names — prevents crash when CSV has two columns with same header
    _seen = {}
    _new_cols = []
    for c in df.columns:
        if c in _seen:
            _seen[c] += 1
            _new_cols.append(f"{c}__{_seen[c]}")
        else:
            _seen[c] = 0
            _new_cols.append(c)
    df.columns = _new_cols
    df = df.dropna(how="all", axis=1).dropna(how="all", axis=0).reset_index(drop=True)

    if TEST_MODE:
        df = df.head(TEST_MODE)
        print(f"✅ Loaded {len(df):,} records (TEST MODE — {TEST_MODE} rows only)")
    else:
        print(f"✅ Loaded {len(df):,} records (full file)")

    # ── SMART CONTENT-BASED COLUMN DETECTION ────────────────────
    # Classifies every column by sampling content, then uses column
    # name only as a tiebreaker. Works on any format.
    US_STATES = {"AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
                 "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
                 "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
                 "VA","WA","WV","WI","WY","DC"}
    BUSINESS_RX = re.compile(
        r'\b(llc|inc|corp|ltd|co\.|company|enterprises|group|services|solutions|consulting|'
        r'trucking|construction|plumbing|electric|mechanical|dental|medical|clinic|salon|'
        r'repair|painting|roofing|cleaning|logistics|staffing|restaurant|hotel|motel|'
        r'bakery|florist|landscaping|insurance|realty|agency|studio|academy|institute)\b', re.I)
    STREET_RX = re.compile(r'^\d+\s+\w+|\b(street|st\b|ave|avenue|blvd|road|rd\b|drive|dr\b|lane|ln\b|court|ct\b|way)\b', re.I)
    STATUS_WORDS = {"verified","invalid","unknown","unverified","disposable","spamtrap",
                    "mismatch","valid","bad","good","ok","deliverable","undeliverable","catchall"}
    FLAG_VALUES  = {"y","n","yes","no","true","false","0","1","t","f"}

    def score_column(colname):
        """Return dict of {field: score} for every field this column could match."""
        sample = df[colname].dropna().astype(str).str.strip()
        sample = sample[sample != ""]
        if len(sample) == 0:
            return {}
        s = sample.head(50)
        name = colname.lower().strip()
        scores = {}

        # Phone: 10-13 digits after stripping non-numeric
        ph = s.apply(lambda v: 10 <= len(re.sub(r'\D', '', v.split('.')[0])) <= 13).sum() / len(s)
        if ph >= 0.7: scores["phone"] = ph

        # Email: has @ and .
        em = s.apply(lambda v: "@" in v and "." in v.split("@")[-1] and len(v) >= 5).sum() / len(s)
        if em >= 0.7: scores["email"] = em

        # State: 2 uppercase letters matching US states
        st = s.apply(lambda v: v.upper() in US_STATES).sum() / len(s)
        if st >= 0.7: scores["state"] = st

        # Zip: exactly 5 digits (or 5-4 format)
        zp = s.apply(lambda v: bool(re.match(r'^\d{5}(-\d{4})?$', v.split('.')[0]))).sum() / len(s)
        if zp >= 0.7: scores["zip"] = zp

        # Status words (verified/invalid/etc)
        sw = s.apply(lambda v: v.lower() in STATUS_WORDS).sum() / len(s)
        if sw >= 0.5: scores["status_word"] = sw

        # Flag (Y/N/1/0)
        fl = s.apply(lambda v: v.lower() in FLAG_VALUES).sum() / len(s)
        if fl >= 0.8: scores["flag"] = fl

        # Revenue/amount: numeric with median >= 1000
        try:
            nums = pd.to_numeric(s.str.replace(r'[$,\s]', '', regex=True), errors='coerce').dropna()
            if len(nums) >= len(s) * 0.8 and nums.median() >= 1000:
                scores["money"] = len(nums) / len(s)
        except Exception:
            pass

        # Address: starts with digits + street words
        ad = s.apply(lambda v: bool(STREET_RX.search(v))).sum() / len(s)
        if ad >= 0.5: scores["address"] = ad

        # Company: business keywords
        co = s.apply(lambda v: bool(BUSINESS_RX.search(v))).sum() / len(s)
        if co >= 0.2: scores["company"] = co

        # Name: alpha-only, 2-40 chars, no digits, no @
        nm = s.apply(lambda v: bool(re.match(r"^[A-Za-z][A-Za-z\s\'\-\.]{1,39}$", v))
                     and "@" not in v).sum() / len(s)
        if nm >= 0.7: scores["name"] = nm
        return scores

    # Score every column
    col_scores = {c: score_column(c) for c in df.columns}

    # Pick best column per field using content score + name hint
    def best_for(field, name_hints, score_key=None, excluded=None):
        score_key = score_key or field
        excluded = excluded or set()
        best, best_score = None, 0
        for c, sc in col_scores.items():
            if c in excluded: continue
            if score_key not in sc: continue
            bonus = 0.2 if any(h in c.lower() for h in name_hints) else 0
            total = sc[score_key] + bonus
            if total > best_score:
                best, best_score = c, total
        return best

    def all_for(field, name_hints, score_key=None, excluded=None):
        score_key = score_key or field
        excluded = excluded or set()
        out = []
        for c, sc in col_scores.items():
            if c in excluded: continue
            if score_key in sc:
                bonus = 0.2 if any(h in c.lower() for h in name_hints) else 0
                out.append((c, sc[score_key] + bonus))
        out.sort(key=lambda x: -x[1])
        return [c for c, _ in out]

    used = set()
    col = {}

    # Phones (all matching, first one is primary)
    phone_cols = all_for("phone", ["phone","mobile","cell","tel","contact"])
    used.update(phone_cols)

    # Emails
    email_cols = all_for("email", ["email","e-mail","mail"])
    used.update(email_cols)

    # Addresses
    address_cols = all_for("address", ["address","street","addr"])
    used.update(address_cols)

    # Money column → rename to Annual Revenue (largest-median money column)
    money_cols = [(c, col_scores[c]["money"]) for c in col_scores if "money" in col_scores[c] and c not in used]
    if money_cols:
        # Pick column with highest median value
        best_money = max(money_cols, key=lambda x: df[x[0]].astype(str).str.replace(r'[$,\s]','',regex=True)
                         .pipe(pd.to_numeric, errors='coerce').median())
        revenue_col = best_money[0]
        if revenue_col != "Annual Revenue":
            df = df.rename(columns={revenue_col: "Annual Revenue"})
        used.add("Annual Revenue")
        col["revenue"] = "Annual Revenue"

    # Company (business keyword score + name hint)
    col["company"] = best_for("company", ["company","business","dba","merchant","organization"], excluded=used)
    if col["company"]: used.add(col["company"])

    # Names — use column name hints to distinguish first vs last
    name_candidates = [c for c in col_scores if "name" in col_scores[c] and c not in used]
    def find_name(hints, excluded):
        best, best_score = None, 0
        for c in name_candidates:
            if c in excluded: continue
            n = c.lower()
            if any(h in n for h in hints):
                score = col_scores[c]["name"] + 0.5
                if score > best_score:
                    best, best_score = c, score
        return best

    col["first"] = find_name(["first","fname","given"], used)
    if col["first"]: used.add(col["first"])
    col["last"] = find_name(["last","lname","lastname","surname","family"], used)
    if col["last"]: used.add(col["last"])

    # DNC flag column (flag score + name hint)
    dnc_col = None
    for c in col_scores:
        if c in used: continue
        if "flag" in col_scores[c] and any(h in c.lower() for h in ["dnc","donot","do_not","optout","opt_out"]):
            dnc_col = c; break
    col["dnc"] = dnc_col
    if dnc_col: used.add(dnc_col)

    # Wireless/line-type flag
    wireless_col = None
    for c in col_scores:
        if c in used: continue
        if "flag" in col_scores[c] and any(h in c.lower() for h in ["wireless","cell","mobile","linetype","line_type"]):
            wireless_col = c; break
    col["wireless"] = wireless_col
    if wireless_col: used.add(wireless_col)

    # Email status column (verified/invalid vocab + name hint)
    email_status_col = None
    for c in col_scores:
        if c in used: continue
        if "status_word" in col_scores[c] and any(h in c.lower() for h in ["email","mail"]):
            email_status_col = c; break
    col["email_status"] = email_status_col
    if email_status_col: used.add(email_status_col)

    # Phone status column
    phone_status_col = None
    for c in col_scores:
        if c in used: continue
        if "status_word" in col_scores[c] and any(h in c.lower() for h in ["phone","number","call"]):
            phone_status_col = c; break
    col["phone_status"] = phone_status_col
    if phone_status_col: used.add(phone_status_col)

    # State
    col["state"] = best_for("state", ["state","province","region"], excluded=used)
    if col["state"]: used.add(col["state"])

    # Lead status / source (not critical, skip if missing)
    _status_hints = {"lead status", "status", "leadstatus", "disposition"}
    _source_hints = {"lead source", "source", "leadsource", "campaign"}
    col["status"] = next((c for c in df.columns if c.lower() in _status_hints), None)
    col["source"] = next((c for c in df.columns if c.lower() in _source_hints), None)

    # ── DETECTION REPORT ─────────────────────────────────────────
    print(f"\n🔎 Columns detected:")
    report_fields = [
        ("First Name", col.get("first")),
        ("Last Name",  col.get("last")),
        ("Company",    col.get("company")),
        ("Phone(s)",   ", ".join(phone_cols) if phone_cols else None),
        ("Email(s)",   ", ".join(email_cols) if email_cols else None),
        ("Address(s)", ", ".join(address_cols) if address_cols else None),
        ("State",      col.get("state")),
        ("Revenue",    col.get("revenue")),
        ("DNC Flag",   col.get("dnc")),
        ("Wireless",   col.get("wireless")),
        ("Email Status", col.get("email_status")),
        ("Phone Status", col.get("phone_status")),
        ("Lead Status", col.get("status")),
    ]
    for label, value in report_fields:
        icon = "✅" if value else "  "
        print(f"   {icon} {label:<14} → {value or '—'}")

    # ── VALIDATION — fail loud with helpful message ─────────────
    missing = []
    if not phone_cols and not email_cols:
        missing.append("phone OR email")
    if not col.get("first") and not col.get("company"):
        missing.append("name OR company")

    if missing:
        print("\n" + "=" * 60)
        print("❌  FILE REJECTED — required columns missing")
        print("=" * 60)
        print(f"\nMissing: {', '.join(missing)}\n")
        print("What to do:")
        if "phone OR email" in missing:
            print("  • Add a column with phone numbers (any name: phone, mobile, cell, phone1, etc).")
            print("    Accepted formats: (555) 123-4567, 5551234567, +1 555 123 4567")
            print("  • Or add a column with emails.")
        if "name OR company" in missing:
            print("  • Add a First Name column or a Company/Business Name column.")
        print("\nColumns found in your file:")
        for c in df.columns:
            print(f"    {c}")
        print("=" * 60)
        sys.exit(1)

    # ── DETERMINE SOURCE NAME ────────────────────────────────────
    # Used as the "List" column value (replaces Qualified/Needs Fixing)
    source_name = STEM  # fallback to filename
    if col.get("source"):
        vals = df[col["source"]].dropna().astype(str).str.strip()
        vals = vals[vals.ne("") & vals.ne("nan")]
        if len(vals) > 0:
            source_name = vals.mode().iloc[0]

    # ── CLEAN SINGLE FIELDS ──────────────────────────────────────
    print("\n🔧 Cleaning...")
    df["_company_clean"] = df[col["company"]].apply(lambda x: str(x).strip().title() if pd.notna(x) else "") if col["company"] else ""
    df["_status"]        = df[col["status"]].fillna("") if col["status"] else ""
    df["_dnc_flag"]      = ""

    if col["first"] and col["last"]:
        df["_first_clean"] = df[col["first"]].apply(clean_name)
        df["_last_clean"]  = df[col["last"]].apply(clean_name)
    elif col["first"]:
        splits = df[col["first"]].apply(lambda x: split_full_name(x))
        df["_first_clean"] = splits.apply(lambda x: x[0])
        df["_last_clean"]  = splits.apply(lambda x: x[1])
    else:
        df["_first_clean"] = ""
        df["_last_clean"]  = ""

    df["_fullname"] = (df["_first_clean"] + " " + df["_last_clean"]).str.strip().str.lower()

    for dc in [c for c in df.columns if "date" in c.lower()]:
        df[dc] = df[dc].apply(clean_date)

    for mc in ["Monthly Revenue", "Annual Revenue", "Annual rev", "Requested Amount",
               "Approved Amount", "requested_funding_amount", "Revenue", "Funding Amount"]:
        if mc in df.columns:
            df[mc] = df[mc].apply(clean_money)

    for dnc_col in ["Closed by DNC", "SMS Opted Out", "DNC", "Do Not Call"]:
        if dnc_col in df.columns:
            df.loc[df[dnc_col].astype(str).str.upper().isin(["Y","YES","1","TRUE"]), "_dnc_flag"] = "YES"

    # ── DUPLICATE FLAGGING ───────────────────────────────────────
    print("🔍 Checking for potential duplicates...")
    df["Duplicate_Flag"] = ""

    if phone_cols:
        df["_phone_primary"] = df[phone_cols[0]].apply(clean_phone)
        phone_dup  = df["_phone_primary"].ne("") & df.duplicated(subset=["_phone_primary"], keep=False)
        same_name  = phone_dup & df.duplicated(subset=["_phone_primary", "_fullname"], keep=False) & df["_fullname"].ne("")
        diff_name  = phone_dup & ~same_name
        df.loc[same_name, "Duplicate_Flag"] = f"Yes — same person ({source_name})"
        df.loc[diff_name, "Duplicate_Flag"] = f"Yes — different name ({source_name})"
    else:
        df["_phone_primary"] = ""

    if email_cols:
        df["_email_primary"] = df[email_cols[0]].apply(clean_email)
        email_dup = df["_email_primary"].ne("") & df.duplicated(subset=["_email_primary"], keep=False)
        df.loc[email_dup & (df["Duplicate_Flag"] == ""), "Duplicate_Flag"] = f"Yes — same email ({source_name})"
    else:
        df["_email_primary"] = ""

    if col["first"] and col["last"]:
        name_dup = df["_fullname"].ne("") & df.duplicated(subset=["_fullname"], keep=False)
        df.loc[name_dup & (df["Duplicate_Flag"] == ""), "Duplicate_Flag"] = f"Yes — same name ({source_name})"

    print(f"   Total flagged: {df['Duplicate_Flag'].ne('').sum()}")

    # ── VERIFY ALL PHONES / EMAILS + BUILD OUTPUT COLS ──────────
    print("\n🔬 Verifying leads (this may take a while)...")
    _total_rows = len(df)
    _done_count = [0]
    import threading
    _lock = threading.Lock()

    def _progress():
        filled = int(30 * _done_count[0] / _total_rows) if _total_rows else 30
        bar = "█" * filled + "░" * (30 - filled)
        sys.stderr.write(f'\r  [Verifying] [{bar}] {_done_count[0]}/{_total_rows}')
        sys.stderr.flush()

    def _process_row(args):
        i, row = args
        csv_name = str(row.get("_fullname", "")).strip()
        company  = str(row.get("_company_clean", "")).strip()

        # Collect + pre-score phones by column name before Twilio
        phone_candidates = []
        for pc in phone_cols:
            v = row.get(pc)
            if not v or pd.isna(v): continue
            for part in _split_multi(v):
                if part and part != "nan":
                    phone_candidates.append((part, pc, _phone_col_score(pc)))
        phone_candidates.sort(key=lambda x: -x[2])
        # Drop fax-only entries if non-fax alternatives exist
        non_fax = [p for p in phone_candidates if p[2] > -5]
        if non_fax:
            phone_candidates = non_fax
        raw_phones          = [p[0] for p in phone_candidates[:3]]
        phone_overflow      = phone_candidates[3:]

        # Collect + pre-score emails by column name + domain quality
        email_candidates = []
        for ec in email_cols:
            v = row.get(ec)
            if not v or pd.isna(v): continue
            for part in _split_multi(v):
                if part and part != "nan" and "@" in part:
                    email_candidates.append((part, ec, _email_col_score(ec, part)))
        email_candidates.sort(key=lambda x: -x[2])
        raw_emails     = [e[0] for e in email_candidates[:3]]
        email_overflow = email_candidates[3:]

        biz_result   = verify_business_google(company)
        sos_result   = verify_sos_cobalt(company)
        google_phone = biz_result.get("found_phone", "")

        if google_phone and google_phone not in raw_phones:
            raw_phones.append(google_phone)

        ranked_phones = rank_phones(raw_phones, csv_name) if raw_phones else []
        ranked_emails = rank_emails(raw_emails) if raw_emails else []

        state = ""
        for ac in address_cols:
            v = row.get(ac, "")
            if v and not pd.isna(v) and str(v).strip():
                state = extract_state(str(v))
                if state:
                    break

        sc = {}
        best_phone = ranked_phones[0] if ranked_phones else {}
        best_email = ranked_emails[0] if ranked_emails else {}

        sc["Best Phone"]      = best_phone.get("cleaned", "")
        sc["Phone Status"]    = best_phone.get("label", "")
        sc["Cell/Landline"]   = line_type_to_label(best_phone.get("line_type", ""))
        sc["Full Name Found"] = best_phone.get("caller_name", "")
        sc["Best Email"]      = best_email.get("cleaned", "")
        sc["Email Status"]    = best_email.get("label", "")
        sc["State"]           = state
        sc["_sos_check"]      = sos_result

        for idx, ph in enumerate(ranked_phones[1:3], 2):
            sc[f"SeaCap_Phone_{idx}"]        = ph.get("cleaned", "")
            sc[f"SeaCap_Phone_{idx}_Status"] = ph.get("label", "")
        for idx, em in enumerate(ranked_emails[1:3], 2):
            sc[f"SeaCap_Email_{idx}"]        = em.get("cleaned", "")
            sc[f"SeaCap_Email_{idx}_Status"] = em.get("label", "")

        # Overflow phones/emails (rank 4+) → Notes with labels
        notes_parts = []
        if phone_overflow:
            labels = []
            for v, col_name, score in phone_overflow:
                reason = "fax column" if score <= -5 else "rank 4+"
                labels.append(f"{v} [{reason}]")
            notes_parts.append("Extra phones (not verified): " + ", ".join(labels))
        if email_overflow:
            labels = []
            for v, col_name, score in email_overflow:
                local = v.split("@")[0].lower()
                role  = any(local.startswith(r) for r in ["info","sales","admin","contact","support","hello","office"])
                reason = "role-based" if role else "rank 4+"
                labels.append(f"{v} [{reason}]")
            notes_parts.append("Extra emails (not verified): " + ", ".join(labels))
        sc["Notes"] = " | ".join(notes_parts) if notes_parts else ""

        reasons = []
        if best_phone.get("label") == "MISMATCH":
            reasons.append("Phone belongs to different person")
        if best_phone.get("label") == "VOIP":
            reasons.append("Phone is VoIP")
        if best_email.get("label") == "SPAMTRAP":
            reasons.append("Email is spamtrap")
        if best_email.get("label") == "DISPOSABLE":
            reasons.append("Email is disposable")
        if biz_result.get("status") in ("VERIFIED-INACTIVE", "PERMANENTLY_CLOSED"):
            reasons.append("Business closed/dissolved")
        sc["Flag Reason"] = " | ".join(reasons)

        with _lock:
            _done_count[0] += 1
            _progress()

        return i, sc

    from concurrent.futures import ThreadPoolExecutor, as_completed
    VERIFY_WORKERS = 10

    results_map = {}
    with ThreadPoolExecutor(max_workers=VERIFY_WORKERS) as executor:
        futures = {executor.submit(_process_row, (i, row)): i for i, row in df.iterrows()}
        for f in as_completed(futures):
            idx, sc = f.result()
            results_map[idx] = sc

    seacap_rows = [results_map[i] for i in df.index]

    # Merge seacap cols into df
    seacap_df = pd.DataFrame(seacap_rows, index=df.index)
    sys.stderr.write('\n')
    sys.stderr.flush()

    df_out = pd.concat([seacap_df, df], axis=1)

    # Dedupe columns created by concat (seacap cols take priority — keep first occurrence)
    df_out = df_out.loc[:, ~df_out.columns.duplicated(keep="first")]

    # ── RENAME DETECTED ORIGINAL COLS → STANDARD NAMES ──────────
    rename_map = {}
    if col.get("first")   and col["first"]   != "First Name": rename_map[col["first"]]   = "First Name"
    if col.get("last")    and col["last"]     != "Last Name":  rename_map[col["last"]]    = "Last Name"
    if col.get("company") and col["company"]  != "Company":    rename_map[col["company"]] = "Company"
    if rename_map:
        df_out = df_out.rename(columns=rename_map)

    # ── CLASSIFY + ASSIGN COLUMNS ───────────────────────────────
    # Attach internal cols needed by get_list
    df_out["_sos_check"] = seacap_df["_sos_check"]
    df_out["_dnc_flag"]  = df["_dnc_flag"]
    df_out["_status"]    = df["_status"]

    # Run classification — store separately so it survives column cleanup
    _lt = df_out.apply(get_list, axis=1)

    df_out["List"] = source_name  # what Adam sees — source name not "Qualified"

    # ── FIX REASON ───────────────────────────────────────────────
    def get_fix_reason(row, list_type):
        if list_type != "Needs Fixing":
            return ""
        reasons = []
        if not str(row.get("Best Phone", "")).strip():
            reasons.append("Missing phone")
        if not str(row.get("Best Email", "")).strip():
            reasons.append("Missing email")
        if str(row.get("Phone Status", "")) == "MISMATCH":
            reasons.append("Phone belongs to different person")
        if str(row.get("_sos_check", "")).lower() == "inactive":
            reasons.append("SOS inactive")
        if str(row.get("Email Status", "")) in ["INVALID", "SPAMTRAP", "DISPOSABLE"]:
            reasons.append("Bad email")
        return ", ".join(reasons) if reasons else "Incomplete data"

    df_out["Fix_Reason"] = [get_fix_reason(df_out.iloc[i], _lt.iloc[i]) for i in range(len(df_out))]

    def get_action_to_fix(row, list_type):
        if list_type != "Needs Fixing":
            return ""
        phone_label = str(row.get("Phone Status", "")).upper()
        email_label = str(row.get("Email Status", "")).upper()
        phone       = str(row.get("Best Phone", "")).strip()
        email       = str(row.get("Best Email", "")).strip()
        first       = str(row.get("First Name", "")).strip()
        last        = str(row.get("Last Name", "")).strip()
        company     = str(row.get("Company", "")).strip()

        if phone_label == "MISMATCH" and not email:
            return "Find correct phone number or email address"
        if phone_label == "MISMATCH" and email_label == "INVALID":
            return "Phone belongs to someone else — correct email address"
        if phone_label == "VERIFIED-NO-NAME" and email_label == "INVALID":
            if not _email_name_match(email, first, last, company):
                return "Correct email address — no name match found"
        if phone_label == "FORMAT-BAD" and email_label == "INVALID":
            return "Fix phone number format and correct email"
        if phone_label == "FORMAT-BAD":
            return "Fix phone number format"
        if not phone and email_label == "INVALID":
            return "Find phone number — email invalid"
        if not phone and not email:
            return "Find phone number and email address"
        return "Verify contact details"

    df_out["Action to Fix"] = [get_action_to_fix(df_out.iloc[i], _lt.iloc[i]) for i in range(len(df_out))]

    # Add Status + Adam List based on list type
    df_out["Status"]    = _lt.map(lambda lt: _LIST_META.get(lt, ("", ""))[0])
    df_out["Adam List"] = _lt.map(lambda lt: _LIST_META.get(lt, ("", ""))[1])
    df_out["Priority"]  = df_out.apply(get_priority, axis=1)

    # Drop internal cols before column ordering
    df_out = df_out.drop(columns=["_dnc_flag", "_status", "_sos_check"], errors="ignore")

    # Drop alt phone/email cols that are fully empty (CSV had no alts)
    alt_cols = [c for c in df_out.columns if c.startswith("SeaCap_Phone_") or c.startswith("SeaCap_Email_")]
    empty_alts = [c for c in alt_cols if df_out[c].replace("", pd.NA).isna().all()]
    df_out = df_out.drop(columns=empty_alts, errors="ignore")

    # Strip illegal chars
    for c in df_out.select_dtypes(include=["object"]).columns:
        df_out[c] = df_out[c].apply(strip_illegal)

    # ── APPLY FIXED COLUMN ORDER ─────────────────────────────────
    alt_cols_remaining = [c for c in df_out.columns if c.startswith("SeaCap_Phone_") or c.startswith("SeaCap_Email_")]
    ordered_cols = [c for c in FIXED_COLS if c in df_out.columns] + alt_cols_remaining

    # ── PRESERVE UNMAPPED SOURCE DATA → Notes ────────────────────
    # Any original column not already in the fixed schema goes into Notes
    ordered_set  = set(ordered_cols)
    unmapped_cols = [c for c in df_out.columns
                     if c not in ordered_set and not str(c).startswith("_")]
    if unmapped_cols:
        def _enrich_notes(row):
            parts = []
            existing = str(row.get("Notes", "")).strip()
            if existing:
                parts.append(existing)
            src = []
            for c in unmapped_cols:
                v = row.get(c, "")
                v_str = str(v).strip()
                if pd.notna(v) and v_str and v_str.lower() not in ("nan", "none", ""):
                    src.append(f"{c}: {v_str}")
            if src:
                parts.append("Source fields: " + " | ".join(src))
            return " || ".join(parts)
        df_out["Notes"] = df_out.apply(_enrich_notes, axis=1)

    df_out = df_out[ordered_cols]

    # ── SPLIT INTO 4 LISTS (using _lt series) ────────────────────

    df_qualified = df_out[_lt == "Qualified"].copy()
    df_fixing    = df_out[_lt == "Needs Fixing"].copy()
    # "Action to Fix" already computed in df_out — carried into df_fixing
    df_dnc       = df_out[_lt == "DNC"].copy()
    df_funded    = df_out[_lt == "Funded"].copy()
    dup_col      = df_out["Duplicate_Flag"] if "Duplicate_Flag" in df_out.columns else pd.Series("", index=df_out.index)
    df_flagged   = df_out[dup_col.ne("")].copy()

    # ── SORT BY QUALITY (highest first) ─────────────────────────
    for frame in [df_qualified, df_fixing, df_dnc, df_funded]:
        frame["_q"] = frame.apply(quality_score, axis=1)
        frame.sort_values("_q", ascending=False, inplace=True)
        frame.drop(columns=["_q"], inplace=True)

    def save_xlsx(df, path):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")
            ws = writer.sheets["Leads"]
            ws.freeze_panes = "A2"
            header_fill = PatternFill("solid", fgColor="D9D9D9")
            for cell in ws[1]:
                cell.font      = Font(bold=True)
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Per-list column orders
    QUALIFIED_COLS = [
        "First Name", "Last Name", "Company", "Adam List", "Priority",
        "Best Phone", "Phone Status", "Cell/Landline",
        "Best Email", "Email Status", "State",
        "Annual Revenue", "Full Name Found", "Duplicate_Flag",
        "Status", "List",
    ]
    FIXING_COLS = [
        "First Name", "Last Name", "Company", "Adam List", "Flag Reason",
        "Best Phone", "Phone Status", "Cell/Landline",
        "Best Email", "Email Status", "State",
        "Annual Revenue", "Full Name Found", "List", "Duplicate_Flag",
        "Status", "Fix_Reason", "Action to Fix",
    ]

    def _reorder(df, cols):
        keep = [c for c in cols if c in df.columns]
        return df[keep]

    print("💾 Saving...")
    save_xlsx(df_out,       OUTPUT_DIR / f"{STEM}_all.xlsx")
    if len(df_qualified): save_xlsx(_reorder(df_qualified, QUALIFIED_COLS), OUTPUT_DIR / f"{STEM}_list1_qualified.xlsx")
    if len(df_fixing):    save_xlsx(_reorder(df_fixing,    FIXING_COLS),    OUTPUT_DIR / f"{STEM}_list2_needs_fixing.xlsx")
    if len(df_dnc):       save_xlsx(df_dnc,       OUTPUT_DIR / f"{STEM}_list3_dnc.xlsx")
    if len(df_funded):    save_xlsx(df_funded,    OUTPUT_DIR / f"{STEM}_list4_funded.xlsx")
    if len(df_flagged):   save_xlsx(df_flagged,   OUTPUT_DIR / f"{STEM}_flagged_for_review.xlsx")

    print("\n" + "="*55)
    print("✅  SUMMARY")
    print("="*55)
    print(f"📊 Total              : {len(df_out):,}")
    print(f"✅ List 1 Qualified   : {len(df_qualified):,}")
    print(f"🔧 List 2 Needs Fix   : {len(df_fixing):,}")
    print(f"🚫 List 3 DNC         : {len(df_dnc):,}")
    print(f"💰 List 4 Funded      : {len(df_funded):,}")
    print(f"⚠️  Flagged for review : {len(df_flagged):,}  ← rep decides")
    print(f"\n📁 Output: {OUTPUT_DIR}")
    print("="*55)

    import json
    summary = {
        "total_leads": len(df_out),
        "list1_qualified": len(df_qualified),
        "list2_needs_fixing": len(df_fixing),
        "list3_dnc": len(df_dnc),
        "list4_funded": len(df_funded),
        "flagged_duplicates": len(df_flagged),
    }
    print(f"SUMMARY_JSON:{json.dumps(summary)}")

if __name__ == "__main__":
    main()
